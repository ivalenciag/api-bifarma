import ast
import csv
import hashlib
import io
import json
import os
import re
import subprocess
import sys
import threading
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, Header, HTTPException, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


load_dotenv()


APP_VERSION = "2.7.0"
BASE_URL = "http://www.datacenter.bifarma.es/bifarma"
LOGIN_URL = f"{BASE_URL}/LoginForm.aspx?ReturnUrl=%2fbifarma"
REQUEST_TIMEOUT = (10, 45)
SESSION_TTL_SECONDS = 25 * 60
DOWNLOAD_TTL_SECONDS = 30 * 60

app = FastAPI(
    title="Bifarma Connector API",
    description=(
        "API protegida con X-API-Key para iniciar sesión en Bifarma, listar "
        "informes, extraer datos y exportarlos en CSV, XLSX o PDF."
    ),
    version=APP_VERSION,
)

_download_lock = threading.RLock()
_download_store: Dict[str, Dict[str, Any]] = {}


@dataclass
class ReportInfo:
    code: str
    title: str
    report_type: Optional[int]
    icon: Optional[str]
    role: Optional[str]
    index: int
    url: str


class BifarmaError(Exception):
    pass


def require_api_key(x_api_key: Optional[str] = Header(default=None, alias="x-api-key")) -> None:
    expected_api_key = os.getenv("API_KEY")
    if not expected_api_key:
        raise HTTPException(status_code=500, detail="API_KEY no está configurada en el servidor.")

    if not x_api_key or x_api_key != expected_api_key:
        raise HTTPException(status_code=401, detail="API key no válida o ausente.")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_code(code: str) -> str:
    return clean_text(code).upper()


def normalize_key(value: str) -> str:
    value = clean_text(value).lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "ñ": "n",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return re.sub(r"[^a-z0-9]+", "", value)


def parse_js_value(raw_value: str) -> Any:
    raw_value = raw_value.strip()
    if not raw_value:
        return ""

    if raw_value[0] in {"'", '"'}:
        try:
            return ast.literal_eval(raw_value)
        except Exception:
            return raw_value.strip("'\"")

    try:
        return int(raw_value)
    except ValueError:
        return raw_value


def now_timestamp() -> float:
    return datetime.utcnow().timestamp()


def safe_filename(value: str) -> str:
    value = normalize_code(value)
    return re.sub(r"[^A-Z0-9_.-]+", "_", value) or "report"


def cleanup_download_store_locked() -> None:
    current = now_timestamp()
    expired_ids = [
        file_id
        for file_id, record in _download_store.items()
        if float(record.get("expires_at_ts", 0)) <= current
    ]
    for file_id in expired_ids:
        _download_store.pop(file_id, None)


def store_download_file(content: bytes, filename: str, media_type: str, export_mode: str) -> Dict[str, Any]:
    if not content:
        raise BifarmaError("No se puede crear enlace de descarga porque el archivo esta vacio.")

    file_id = uuid.uuid4().hex
    token = uuid.uuid4().hex
    expires_at_ts = now_timestamp() + DOWNLOAD_TTL_SECONDS

    with _download_lock:
        cleanup_download_store_locked()
        _download_store[file_id] = {
            "content": content,
            "filename": filename,
            "media_type": media_type,
            "token": token,
            "created_at_ts": now_timestamp(),
            "expires_at_ts": expires_at_ts,
            "export_mode": export_mode,
        }

    return {
        "file_id": file_id,
        "token": token,
        "filename": filename,
        "media_type": media_type,
        "size_bytes": len(content),
        "expires_at": datetime.utcfromtimestamp(expires_at_ts).isoformat() + "Z",
        "export_mode": export_mode,
    }


def get_download_file(file_id: str, token: str) -> Dict[str, Any]:
    with _download_lock:
        cleanup_download_store_locked()
        record = _download_store.get(file_id)
        if not record:
            raise HTTPException(status_code=404, detail="El enlace de descarga no existe o ha caducado.")
        if not token or token != record.get("token"):
            raise HTTPException(status_code=403, detail="Token de descarga no valido.")
        return record


def format_bifarma_date(value: Optional[str]) -> Optional[str]:
    if not value:
        return None

    value = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass

    return value


def is_yyyymm(value: Optional[str]) -> bool:
    """True si el valor es un mes en formato YYYYMM (ej. 202606)."""
    return bool(value and re.fullmatch(r"20\d{4}", value.strip()))


def parse_ddmmyyyy(value: Optional[str]) -> Optional[Tuple[int, int, int]]:
    """dd/mm/aaaa -> (anio, mes, dia). Devuelve None si no encaja."""
    if not value:
        return None
    m = re.fullmatch(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*", value)
    if not m:
        return None
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        datetime(year, month, day)
    except ValueError:
        return None
    return year, month, day


def looks_like_login_page(html: str) -> bool:
    html_lower = html.lower()
    return "loginform.aspx" in html_lower or 'type="password"' in html_lower or "type='password'" in html_lower


def make_document_url(role: Optional[str], code: str) -> str:
    if not role:
        raise BifarmaError(f"No se ha podido detectar BIRoleId para el informe {code}.")

    return (
        f"{BASE_URL}/Private/BIDocumentManager.aspx"
        f"?BIRoleId={quote(role, safe='')}&BIDocId={quote(code, safe='')}"
    )


def parse_report_menu(html: str) -> List[ReportInfo]:
    assignment_re = re.compile(
        r"(?P<var>dI|d|T|D|r)\s*\[\s*(?P<idx>\d+)\s*\]\s*=\s*"
        r"(?P<value>\"(?:\\.|[^\"])*\"|'(?:\\.|[^'])*'|-?\d+)",
        re.IGNORECASE,
    )

    by_index: Dict[int, Dict[str, Any]] = {}
    role_positions: List[Tuple[int, str]] = []
    code_positions: Dict[int, int] = {}

    for match in assignment_re.finditer(html):
        var_name = match.group("var")
        index = int(match.group("idx"))
        value = parse_js_value(match.group("value"))

        if var_name == "r":
            role_positions.append((match.start(), clean_text(value)))
            continue

        by_index.setdefault(index, {})[var_name] = value
        if var_name == "d":
            code_positions[index] = match.start()

    reports: List[ReportInfo] = []
    for index, values in sorted(by_index.items()):
        code = clean_text(values.get("d"))
        if not code:
            continue

        role = None
        code_position = code_positions.get(index, 0)
        previous_roles = [item for item in role_positions if item[0] < code_position and item[1]]
        if previous_roles:
            role = previous_roles[-1][1]

        title = clean_text(values.get("D")) or code
        report_type = values.get("T")
        if not isinstance(report_type, int):
            try:
                report_type = int(report_type)
            except Exception:
                report_type = None

        reports.append(
            ReportInfo(
                code=normalize_code(code),
                title=title,
                report_type=report_type,
                icon=clean_text(values.get("dI")) or None,
                role=role,
                index=index,
                url=make_document_url(role, normalize_code(code)) if role else "",
            )
        )

    return reports


def report_to_dict(report: ReportInfo) -> Dict[str, Any]:
    return {
        "code": report.code,
        "title": report.title,
        "type": report.report_type,
        "icon": report.icon,
        "role": report.role,
        "index": report.index,
        "url": report.url,
    }


def collect_form_fields(form: BeautifulSoup) -> Dict[str, str]:
    payload: Dict[str, str] = {}

    for input_tag in form.find_all("input"):
        name = input_tag.get("name")
        if not name:
            continue

        input_type = clean_text(input_tag.get("type")).lower()
        if input_type in {"button", "submit", "image", "reset", "file"}:
            continue

        if input_type in {"checkbox", "radio"} and not input_tag.has_attr("checked"):
            continue

        payload[name] = input_tag.get("value", "")

    for textarea in form.find_all("textarea"):
        name = textarea.get("name")
        if name:
            payload[name] = textarea.text or ""

    for select in form.find_all("select"):
        name = select.get("name")
        if not name:
            continue
        selected = select.find("option", selected=True) or select.find("option")
        payload[name] = selected.get("value", selected.text if selected else "") if selected else ""

    return payload


def choose_login_form(soup: BeautifulSoup) -> Optional[BeautifulSoup]:
    forms = soup.find_all("form")
    if not forms:
        return None

    for form in forms:
        if form.find("input", {"type": re.compile("password", re.I)}):
            return form

    return forms[0]


def find_login_field(form: BeautifulSoup, kind: str) -> Optional[str]:
    inputs = form.find_all("input")

    if kind == "password":
        for input_tag in inputs:
            if clean_text(input_tag.get("type")).lower() == "password" and input_tag.get("name"):
                return input_tag.get("name")

    user_hints = ("user", "usuario", "login", "email", "mail", "name", "cuenta")
    for input_tag in inputs:
        name = input_tag.get("name")
        if not name:
            continue

        input_type = clean_text(input_tag.get("type")).lower()
        key = normalize_key(f"{name} {input_tag.get('id', '')}")

        if kind == "username" and input_type in {"text", "email", ""}:
            if any(hint in key for hint in user_hints):
                return name

    if kind == "username":
        for input_tag in inputs:
            name = input_tag.get("name")
            input_type = clean_text(input_tag.get("type")).lower()
            if name and input_type in {"text", "email", ""}:
                return name

    return None


def choose_submit_button(form: BeautifulSoup) -> Optional[Tuple[str, str]]:
    submit_hints = ("entrar", "login", "acceder", "aceptar", "consultar", "buscar", "ver", "generar")
    buttons = form.find_all(["input", "button"])

    for button in buttons:
        button_type = clean_text(button.get("type")).lower()
        if button_type not in {"submit", "button", "image", ""}:
            continue

        name = button.get("name")
        if not name:
            continue

        button_text = normalize_key(f"{button.get('value', '')} {button.text} {button.get('id', '')} {name}")
        if any(hint in button_text for hint in submit_hints):
            return name, button.get("value", clean_text(button.text))

    for button in buttons:
        button_type = clean_text(button.get("type")).lower()
        name = button.get("name")
        if name and button_type in {"submit", "button", "image", ""}:
            return name, button.get("value", clean_text(button.text))

    return None


def extract_html_title(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    if soup.title:
        return clean_text(soup.title.get_text(" "))
    heading = soup.find(["h1", "h2", "h3"])
    return clean_text(heading.get_text(" ")) if heading else ""


def form_field_summary(html: str) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    forms_summary: List[Dict[str, Any]] = []

    for form_index, form in enumerate(soup.find_all("form"), start=1):
        fields: List[Dict[str, Any]] = []

        for tag in form.find_all(["input", "select", "textarea", "button"]):
            name = tag.get("name")
            field_id = tag.get("id")
            field_type = clean_text(tag.get("type") or tag.name).lower()
            label_text = ""

            if field_id:
                label = soup.find("label", attrs={"for": field_id})
                if label:
                    label_text = clean_text(label.get_text(" "))

            fields.append(
                {
                    "tag": tag.name,
                    "type": field_type,
                    "name": name,
                    "id": field_id,
                    "label": label_text,
                    "looks_like_date": looks_like_date_field(name, field_id, label_text),
                }
            )

        forms_summary.append(
            {
                "form_index": form_index,
                "method": clean_text(form.get("method") or "get").lower(),
                "action": form.get("action") or "",
                "field_count": len(fields),
                "fields": fields,
            }
        )

    return forms_summary


def looks_like_date_field(*parts: Optional[str]) -> bool:
    key = normalize_key(" ".join(part or "" for part in parts))
    hints = (
        "fecha",
        "desde",
        "hasta",
        "inicio",
        "fin",
        "date",
        "start",
        "end",
        "from",
        "to",
        "fec",
    )
    return any(hint in key for hint in hints)


def pick_report_form(soup: BeautifulSoup) -> Optional[BeautifulSoup]:
    forms = soup.find_all("form")
    if not forms:
        return None

    for form in forms:
        text = normalize_key(form.get_text(" "))
        field_names = normalize_key(" ".join(tag.get("name", "") + " " + tag.get("id", "") for tag in form.find_all(True)))
        if any(hint in f"{text} {field_names}" for hint in ("fecha", "desde", "hasta", "consultar", "buscar")):
            return form

    for form in forms:
        if form.find("input", attrs={"name": re.compile("__VIEWSTATE", re.I)}):
            return form

    return forms[0]


def find_matching_field(payload: Dict[str, str], fields: List[Dict[str, Any]], hints: Tuple[str, ...]) -> Optional[str]:
    best_name = None
    best_score = 0

    for field in fields:
        name = field.get("name")
        if not name or name not in payload:
            continue

        field_type = clean_text(field.get("type")).lower()
        if field_type in {"submit", "button", "image", "reset", "file"}:
            continue

        key = normalize_key(f"{name} {field.get('id', '')} {field.get('label', '')}")
        score = 0
        for hint in hints:
            if hint in key:
                score += 10

        if "fecha" in key or "date" in key or "fec" in key:
            score += 4
        if "hidden" in field_type and score < 10:
            score -= 5

        if score > best_score:
            best_name = name
            best_score = score

    return best_name


def parse_filter_json(filters: Optional[str]) -> Dict[str, str]:
    if not filters:
        return {}

    try:
        parsed = json.loads(filters)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"El parámetro filters no es JSON válido: {exc.msg}")

    if not isinstance(parsed, dict):
        raise HTTPException(status_code=400, detail="El parámetro filters debe ser un objeto JSON.")

    return {clean_text(key): clean_text(value) for key, value in parsed.items() if clean_text(key)}


def build_requested_filters(
    request: Request,
    start_date: Optional[str],
    end_date: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    filters: Optional[str],
) -> Dict[str, str]:
    reserved = {
        "format",
        "mode",
        "max_rows",
        "all_pages",
        "max_pages",
        "start_date",
        "end_date",
        "date_from",
        "date_to",
        "filters",
    }
    requested = parse_filter_json(filters)

    query_params = dict(request.query_params)
    start_value = (
        date_from
        or start_date
        or query_params.get("fecha_inicio")
        or query_params.get("fecha_desde")
        or query_params.get("desde")
    )
    end_value = (
        date_to
        or end_date
        or query_params.get("fecha_fin")
        or query_params.get("fecha_hasta")
        or query_params.get("hasta")
    )

    if start_value:
        if is_yyyymm(start_value):
            requested["__month_from__"] = start_value.strip()
        else:
            requested["__date_from__"] = format_bifarma_date(start_value) or start_value
    if end_value:
        if is_yyyymm(end_value):
            requested["__month_to__"] = end_value.strip()
        else:
            requested["__date_to__"] = format_bifarma_date(end_value) or end_value

    for key, value in request.query_params.multi_items():
        if key not in reserved and key not in requested and clean_text(value):
            requested[key] = clean_text(value)

    return requested


def is_numberish(value: str) -> bool:
    value = clean_text(value)
    if not value:
        return False

    cleaned = value.replace("€", "").replace("%", "").replace(" ", "")
    cleaned = cleaned.replace(".", "").replace(",", ".")
    cleaned = cleaned.replace("+", "").replace("-", "", 1)

    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def normalize_cell(cell: BeautifulSoup) -> str:
    for unwanted in cell.find_all(["script", "style"]):
        unwanted.decompose()
    return clean_text(cell.get_text(" "))


def dedupe_columns(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    columns: List[str] = []

    for position, header in enumerate(headers, start=1):
        name = clean_text(header) or f"Columna {position}"
        if name not in seen:
            seen[name] = 1
            columns.append(name)
            continue

        seen[name] += 1
        columns.append(f"{name} ({seen[name]})")

    return columns


def header_score(row: List[str]) -> int:
    if not row:
        return -100

    non_empty = [cell for cell in row if cell]
    if len(non_empty) < 2:
        return -50

    score = len(non_empty)
    numeric_count = sum(1 for cell in non_empty if is_numberish(cell))
    score -= numeric_count * 2

    joined = normalize_key(" ".join(non_empty))
    business_hints = (
        "producto",
        "laboratorio",
        "importe",
        "margen",
        "stock",
        "uds",
        "unidades",
        "cliente",
        "cn",
        "ventas",
    )
    score += sum(5 for hint in business_hints if hint in joined)

    if len(" ".join(non_empty)) > 300:
        score -= 10

    return score


def extract_tables(html: str, max_rows: int) -> Dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    candidates: List[Dict[str, Any]] = []

    for table_index, table in enumerate(soup.find_all("table"), start=1):
        raw_rows: List[List[str]] = []
        header_row_indexes = set()

        for row_index, tr in enumerate(table.find_all("tr")):
            cells = tr.find_all(["th", "td"])
            row = [normalize_cell(cell) for cell in cells]
            if any(row):
                raw_rows.append(row)
                if tr.find_all("th"):
                    header_row_indexes.add(len(raw_rows) - 1)

        if len(raw_rows) < 2:
            continue

        max_width = max(len(row) for row in raw_rows)
        normalized_rows = [row + [""] * (max_width - len(row)) for row in raw_rows]

        if header_row_indexes:
            header_index = min(header_row_indexes)
        else:
            scored_rows = [(idx, header_score(row)) for idx, row in enumerate(normalized_rows[:8])]
            header_index = max(scored_rows, key=lambda item: item[1])[0]

        headers = dedupe_columns(normalized_rows[header_index])
        data_rows = normalized_rows[header_index + 1 :]
        data_rows = [row for row in data_rows if any(clean_text(cell) for cell in row)]

        if not data_rows or len(headers) < 2:
            continue

        rows_as_dicts = [
            {headers[column_index]: clean_text(value) for column_index, value in enumerate(row[: len(headers)])}
            for row in data_rows
        ]

        known_header_bonus = header_score(headers)
        table_score = (len(rows_as_dicts) * len(headers)) + known_header_bonus
        candidates.append(
            {
                "table_index": table_index,
                "score": table_score,
                "columns": headers,
                "rows": rows_as_dicts,
                "total_rows": len(rows_as_dicts),
            }
        )

    if not candidates:
        return {
            "table_index": None,
            "columns": ["texto"],
            "rows": [{"texto": clean_text(soup.get_text(" "))[:5000]}],
            "total_rows": 1,
            "tables_detected": 0,
        }

    best = max(candidates, key=lambda item: item["score"])
    return {
        "table_index": best["table_index"],
        "columns": best["columns"],
        "rows": best["rows"][:max_rows],
        "total_rows": best["total_rows"],
        "tables_detected": len(candidates),
    }


POSTBACK_RE = re.compile(
    r"__doPostBack\(\s*['\"](?P<target>[^'\"]*)['\"]\s*,\s*['\"](?P<argument>[^'\"]*)['\"]\s*\)",
    re.IGNORECASE,
)


def parse_spanish_int(value: str) -> Optional[int]:
    digits = re.sub(r"[^\d]", "", clean_text(value))
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def extract_pagination_hint(html: str) -> Dict[str, Optional[int]]:
    soup = BeautifulSoup(html, "html.parser")
    text = clean_text(soup.get_text(" "))
    hint: Dict[str, Optional[int]] = {
        "total_rows_detected": None,
        "total_pages_detected": None,
        "current_page_detected": None,
    }

    rows_match = re.search(
        r"([\d\.\,]+)\s+(?:art[ií]culos|registros|productos|filas)",
        text,
        re.IGNORECASE,
    )
    if rows_match:
        hint["total_rows_detected"] = parse_spanish_int(rows_match.group(1))

    pages_match = re.search(r"([\d\.\,]+)\s+p[aá]ginas", text, re.IGNORECASE)
    if pages_match:
        hint["total_pages_detected"] = parse_spanish_int(pages_match.group(1))

    current_match = re.search(
        r"p[aá]gina\s+([\d\.\,]+)\s+de\s+([\d\.\,]+)",
        text,
        re.IGNORECASE,
    )
    if current_match:
        hint["current_page_detected"] = parse_spanish_int(current_match.group(1))
        hint["total_pages_detected"] = parse_spanish_int(current_match.group(2))

    return hint


def find_postback_events(html: str) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    events: List[Dict[str, Any]] = []
    seen = set()

    for tag in soup.find_all(["a", "input", "button"]):
        fragments = [
            tag.get("href") or "",
            tag.get("onclick") or "",
            tag.get("onmousedown") or "",
        ]
        text = clean_text(
            tag.get_text(" ")
            or tag.get("value")
            or tag.get("title")
            or tag.get("alt")
            or tag.get("aria-label")
            or ""
        )

        for fragment in fragments:
            for match in POSTBACK_RE.finditer(fragment):
                target = match.group("target").replace("\\'", "'").replace('\\"', '"')
                argument = match.group("argument").replace("\\'", "'").replace('\\"', '"')
                key = (target, argument, text)
                if key in seen:
                    continue
                seen.add(key)

                combined = normalize_key(f"{target} {argument} {text} {fragment}")
                is_pagination = (
                    "page" in normalize_key(argument)
                    or "pagina" in combined
                    or "siguiente" in combined
                    or "next" in combined
                    or text in {">", ">>", "»"}
                )
                is_previous = (
                    "page$prev" in argument.lower()
                    or "page$first" in argument.lower()
                    or "anterior" in combined
                    or "previous" in combined
                    or text in {"<", "<<", "«"}
                )

                events.append(
                    {
                        "target": target,
                        "argument": argument,
                        "text": text,
                        "is_pagination": is_pagination,
                        "is_previous": is_previous,
                    }
                )

    return events


def find_next_pagination_event(html: str, current_page: int) -> Optional[Dict[str, Any]]:
    events = [
        event
        for event in find_postback_events(html)
        if event["is_pagination"] and not event["is_previous"]
    ]

    next_page_argument = f"page${current_page + 1}".lower()
    for event in events:
        if event["argument"].lower() == next_page_argument:
            return event

    for event in events:
        if "page$next" in event["argument"].lower():
            return event

    for event in events:
        combined = normalize_key(f"{event['text']} {event['argument']}")
        if "siguiente" in combined or "next" in combined or event["text"] in {">", ">>", "»"}:
            return event

    numeric_events: List[Tuple[int, Dict[str, Any]]] = []
    for event in events:
        text_page = parse_spanish_int(event["text"])
        argument_match = re.search(r"page\$(\d+)", event["argument"], re.IGNORECASE)
        argument_page = int(argument_match.group(1)) if argument_match else None
        page_number = argument_page or text_page
        if page_number and page_number > current_page:
            numeric_events.append((page_number, event))

    if numeric_events:
        return sorted(numeric_events, key=lambda item: item[0])[0][1]

    return None


def rows_signature(rows: List[Dict[str, Any]]) -> str:
    if not rows:
        return ""

    sample = rows[:2] + rows[-2:]
    payload = json.dumps(sample, sort_keys=True, ensure_ascii=False)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def merge_columns(existing: List[str], new_columns: List[str]) -> List[str]:
    merged = list(existing)
    for column in new_columns:
        if column not in merged:
            merged.append(column)
    return merged


def normalize_rows_to_columns(rows: List[Dict[str, Any]], columns: List[str]) -> List[Dict[str, Any]]:
    return [{column: clean_text(row.get(column, "")) for column in columns} for row in rows]


EXPORT_KEYWORDS = (
    "excel",
    "xls",
    "xlsx",
    "csv",
    "export",
    "exportar",
    "descargar",
    "download",
    "aspxgridviewexporter",
    "exporttoxlsx",
    "exporttoxls",
    "exporttocsv",
)


def is_export_candidate_text(value: str, preferred_format: str) -> bool:
    key = normalize_key(value)
    if not key:
        return False

    if preferred_format == "csv":
        preferred = ("csv", "exporttocsv")
    else:
        preferred = ("excel", "xls", "xlsx", "exporttoxlsx", "exporttoxls")

    has_export_word = any(word in key for word in EXPORT_KEYWORDS)
    has_preferred_word = any(word in key for word in preferred)
    return has_export_word and (has_preferred_word or "export" in key or "exportar" in key)


def score_export_candidate(value: str, preferred_format: str) -> int:
    key = normalize_key(value)
    score = 0

    for word in EXPORT_KEYWORDS:
        if word in key:
            score += 10

    if preferred_format == "csv":
        if "csv" in key:
            score += 40
        if "excel" in key or "xlsx" in key or "xls" in key:
            score -= 10
    else:
        if "xlsx" in key or "exporttoxlsx" in key:
            score += 45
        if "excel" in key or "xls" in key or "exporttoxls" in key:
            score += 35
        if "csv" in key:
            score -= 10

    if "pdf" in key:
        score -= 30

    return score


def find_native_export_events(html: str, preferred_format: str = "xlsx") -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    events: List[Dict[str, Any]] = []
    seen = set()

    for tag in soup.find_all(["a", "input", "button", "div", "span", "td", "li"]):
        attrs_text = " ".join(
            clean_text(tag.get(attr))
            for attr in ("id", "name", "value", "title", "alt", "aria-label", "class")
            if tag.get(attr)
        )
        text = clean_text(tag.get_text(" ") or tag.get("value") or tag.get("title") or tag.get("alt") or "")
        fragments = [
            tag.get("href") or "",
            tag.get("onclick") or "",
            tag.get("onmousedown") or "",
            tag.get("data-postback") or "",
            tag.get("data-callback") or "",
        ]
        combined = clean_text(f"{tag.name} {attrs_text} {text} {' '.join(fragments)}")

        if not is_export_candidate_text(combined, preferred_format):
            continue

        href = tag.get("href") or ""
        if href and not href.lower().startswith("javascript:"):
            key = ("direct_url", href)
            if key not in seen:
                seen.add(key)
                events.append(
                    {
                        "kind": "direct_url",
                        "url": href,
                        "text": text,
                        "score": score_export_candidate(combined, preferred_format) + 20,
                    }
                )

        for fragment in fragments:
            for match in POSTBACK_RE.finditer(fragment):
                target = match.group("target").replace("\\'", "'").replace('\\"', '"')
                argument = match.group("argument").replace("\\'", "'").replace('\\"', '"')
                key = ("postback", target, argument)
                if key in seen:
                    continue
                seen.add(key)
                events.append(
                    {
                        "kind": "postback",
                        "target": target,
                        "argument": argument,
                        "text": text,
                        "score": score_export_candidate(combined, preferred_format) + 30,
                    }
                )

        name = tag.get("name")
        if tag.name in {"input", "button"} and name:
            key = ("form_button", name, tag.get("value", ""))
            if key not in seen:
                seen.add(key)
                events.append(
                    {
                        "kind": "form_button",
                        "name": name,
                        "value": tag.get("value", clean_text(tag.get_text(" "))),
                        "text": text,
                        "score": score_export_candidate(combined, preferred_format),
                    }
                )

    for match in POSTBACK_RE.finditer(html):
        target = match.group("target").replace("\\'", "'").replace('\\"', '"')
        argument = match.group("argument").replace("\\'", "'").replace('\\"', '"')
        combined = f"{target} {argument}"
        if not is_export_candidate_text(combined, preferred_format):
            continue
        key = ("postback", target, argument)
        if key in seen:
            continue
        seen.add(key)
        events.append(
            {
                "kind": "postback",
                "target": target,
                "argument": argument,
                "text": "",
                "score": score_export_candidate(combined, preferred_format),
            }
        )

    return sorted(events, key=lambda event: event.get("score", 0), reverse=True)


def extract_export_script_snippets(html: str, limit: int = 12) -> List[Dict[str, str]]:
    snippets: List[Dict[str, str]] = []
    seen = set()

    for keyword in EXPORT_KEYWORDS + ("contextmenu", "popupmenu", "menuitem", "rightclick", "rowcontextmenu"):
        for match in re.finditer(re.escape(keyword), html, re.IGNORECASE):
            start = max(match.start() - 180, 0)
            end = min(match.end() + 260, len(html))
            snippet = clean_text(re.sub(r"<[^>]+>", " ", html[start:end]))
            key = snippet[:160]
            if key in seen:
                continue
            seen.add(key)
            snippets.append({"keyword": keyword, "snippet": snippet[:700]})
            if len(snippets) >= limit:
                return snippets

    return snippets


def response_looks_like_html(content: bytes, content_type: str) -> bool:
    preview = content[:2000].lower()
    return b"<html" in preview or b"<!doctype html" in preview or "text/html" in content_type


def response_looks_like_native_file(response: requests.Response, expected_format: str) -> bool:
    content = response.content or b""
    content_type = clean_text(response.headers.get("content-type")).lower()
    disposition = clean_text(response.headers.get("content-disposition")).lower()
    preview = content[:4000].lower()

    if not content:
        return False

    if response_looks_like_html(content, content_type):
        return False

    if expected_format == "csv":
        if "csv" in content_type or ".csv" in disposition:
            return True
        if b";" in preview or b"," in preview:
            return b"actualizar datos" not in preview and b"cargando" not in preview
        return False

    if content.startswith(b"PK"):
        return True

    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return True

    if "excel" in content_type or "spreadsheet" in content_type:
        return True

    if ".xlsx" in disposition or ".xls" in disposition:
        return True

    return False


def native_media_type(format_value: str, content: bytes) -> str:
    if format_value == "csv":
        return "text/csv; charset=utf-8"
    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return "application/vnd.ms-excel"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def native_filename(code: str, format_value: str, content: bytes) -> str:
    base = safe_filename(code)
    if format_value == "csv":
        return f"{base}.csv"
    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return f"{base}.xls"
    return f"{base}.xlsx"


def playwright_import_status() -> Dict[str, Any]:
    try:
        import playwright  # noqa: F401

        return {"available": True, "error": ""}
    except Exception as exc:
        return {"available": False, "error": clean_text(exc)}


def install_chromium_runtime() -> Dict[str, Any]:
    try:
        completed = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True,
            text=True,
            timeout=180,
        )
        return {
            "ok": completed.returncode == 0,
            "returncode": completed.returncode,
            "stdout_tail": clean_text(completed.stdout[-1200:]),
            "stderr_tail": clean_text(completed.stderr[-1200:]),
        }
    except Exception as exc:
        return {"ok": False, "error": clean_text(exc)}


def chromium_launch_status(auto_install: bool = False) -> Dict[str, Any]:
    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            version = browser.version
            browser.close()
            return {"ok": True, "chromium_version": version, "error": ""}
    except Exception as exc:
        first_error = clean_text(exc)
        install_result = None
        if auto_install:
            install_result = install_chromium_runtime()
            if install_result.get("ok"):
                retry = chromium_launch_status(auto_install=False)
                retry["runtime_install"] = install_result
                retry["first_error"] = first_error
                return retry

        return {
            "ok": False,
            "chromium_version": "",
            "error": first_error,
            "runtime_install": install_result,
            "hint": "En Render usa como Build Command: bash render-build.sh",
        }


def build_csv_bytes(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, delimiter=";", extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def build_xlsx_bytes(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informe"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(columns)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(clean_text(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_pdf_bytes(title: str, columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(clean_text(title) or "Informe Bifarma", styles["Title"]), Spacer(1, 12)]

    limited_columns = columns[:8]
    limited_rows = rows[:80]
    table_data = [limited_columns]
    for row in limited_rows:
        table_data.append([clean_text(row.get(column, ""))[:80] for column in limited_columns])

    pdf_table = Table(table_data, repeatRows=1)
    pdf_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6FA")]),
            ]
        )
    )
    story.append(pdf_table)
    document.build(story)
    output.seek(0)
    return output.getvalue()


class BifarmaClient:
    def __init__(self, username: str, password: str) -> None:
        self.username = username
        self.password = password
        self.lock = threading.RLock()
        self.session = requests.Session()
        self.last_login_at = 0.0
        self.last_html = ""
        self.last_url = ""
        self.menu_cache: Optional[List[ReportInfo]] = None
        self.menu_cache_at = 0.0

    def _new_session(self) -> requests.Session:
        session = requests.Session()
        session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) BifarmaConnector/2.1",
                "Accept-Language": "es-ES,es;q=0.9",
            }
        )
        return session

    def ensure_login(self, force: bool = False) -> Dict[str, Any]:
        with self.lock:
            if not force and self.last_login_at and now_timestamp() - self.last_login_at < SESSION_TTL_SECONDS:
                return {"logged_in": True, "reused_session": True}

            username = self.username
            password = self.password
            if not username or not password:
                raise BifarmaError("Las credenciales de Bifarma son obligatorias.")

            self.session = self._new_session()
            login_page = self.session.get(LOGIN_URL, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            login_page.raise_for_status()

            soup = BeautifulSoup(login_page.text, "html.parser")
            form = choose_login_form(soup)
            if form is None:
                raise BifarmaError("No se ha encontrado el formulario de login de Bifarma.")

            payload = collect_form_fields(form)
            username_field = find_login_field(form, "username")
            password_field = find_login_field(form, "password")

            if not username_field or not password_field:
                raise BifarmaError("No se han podido detectar los campos de usuario y contraseña del login.")

            payload[username_field] = username
            payload[password_field] = password

            submit_button = choose_submit_button(form)
            if submit_button:
                payload[submit_button[0]] = submit_button[1]

            action_url = urljoin(login_page.url, form.get("action") or LOGIN_URL)
            method = clean_text(form.get("method") or "post").lower()

            if method == "get":
                response = self.session.get(action_url, params=payload, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            else:
                response = self.session.post(action_url, data=payload, timeout=REQUEST_TIMEOUT, allow_redirects=True)

            response.raise_for_status()
            self.last_html = response.text
            self.last_url = response.url

            if looks_like_login_page(response.text) and "BIDocumentManager" not in response.text:
                raise BifarmaError("Bifarma ha devuelto de nuevo la pantalla de login. Revisa las credenciales.")

            self.last_login_at = now_timestamp()
            self.menu_cache = None
            self.menu_cache_at = 0.0
            return {"logged_in": True, "reused_session": False, "url": response.url}

    def _fetch_menu_html_locked(self) -> Tuple[str, str]:
        if self.last_html and "d[" in self.last_html and "D[" in self.last_html:
            return self.last_html, self.last_url

        candidate_urls = [
            f"{BASE_URL}/",
            f"{BASE_URL}/Private/",
            f"{BASE_URL}/Private/BIMenu.aspx",
            f"{BASE_URL}/Private/Menu.aspx",
            f"{BASE_URL}/Private/BIReports.aspx",
            f"{BASE_URL}/Private/BIDocumentManager.aspx",
        ]

        checked_pages: List[Tuple[str, str]] = []
        for url in candidate_urls:
            response = self.session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            if response.status_code >= 400:
                continue

            checked_pages.append((response.text, response.url))
            if "d[" in response.text and "D[" in response.text:
                self.last_html = response.text
                self.last_url = response.url
                return response.text, response.url

            soup = BeautifulSoup(response.text, "html.parser")
            frame_urls = []
            for tag in soup.find_all(["frame", "iframe"]):
                src = tag.get("src")
                if src:
                    frame_urls.append(urljoin(response.url, src))

            for frame_url in frame_urls:
                frame_response = self.session.get(frame_url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
                checked_pages.append((frame_response.text, frame_response.url))
                if "d[" in frame_response.text and "D[" in frame_response.text:
                    self.last_html = frame_response.text
                    self.last_url = frame_response.url
                    return frame_response.text, frame_response.url

        for html, url in checked_pages:
            if "BIDocId" in html or "BIRoleId" in html:
                return html, url

        raise BifarmaError("No se ha encontrado el menú privado de informes después del login.")

    def reports_menu(self, force_refresh: bool = False) -> Dict[str, Any]:
        with self.lock:
            self.ensure_login()

            if (
                not force_refresh
                and self.menu_cache is not None
                and now_timestamp() - self.menu_cache_at < SESSION_TTL_SECONDS
            ):
                return {
                    "source_url": self.last_url,
                    "count": len(self.menu_cache),
                    "reports": [report_to_dict(report) for report in self.menu_cache],
                    "cached": True,
                }

            html, source_url = self._fetch_menu_html_locked()
            reports = parse_report_menu(html)
            self.menu_cache = reports
            self.menu_cache_at = now_timestamp()
            self.last_url = source_url

            return {
                "source_url": source_url,
                "count": len(reports),
                "reports": [report_to_dict(report) for report in reports],
                "cached": False,
            }

    def find_report(self, code: str) -> ReportInfo:
        normalized = normalize_code(code)
        menu = self.reports_menu()
        reports = [ReportInfo(**self._dict_to_report_kwargs(report)) for report in menu["reports"]]

        for report in reports:
            if report.code == normalized:
                return report

        partial_matches = [report for report in reports if normalized in report.code or normalized in report.title.upper()]
        if partial_matches:
            return partial_matches[0]

        raise BifarmaError(f"No se ha encontrado el informe {normalized} en el menú de Bifarma.")

    @staticmethod
    def _dict_to_report_kwargs(report: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "code": report["code"],
            "title": report["title"],
            "report_type": report.get("type"),
            "icon": report.get("icon"),
            "role": report.get("role"),
            "index": report.get("index"),
            "url": report.get("url"),
        }

    def open_report_html(self, code: str, requested_filters: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
        with self.lock:
            self.ensure_login()
            report = self.find_report(code)
            url = make_document_url(report.role, report.code)

            response = self.session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            response.raise_for_status()

            html = response.text
            final_url = response.url
            filter_status = {
                "requested": requested_filters or {},
                "applied": {},
                "not_applied": {},
                "submitted": False,
                "message": "No se han solicitado filtros.",
            }

            if requested_filters:
                html, final_url, filter_status = self._apply_report_filters(
                    html=html,
                    current_url=response.url,
                    requested_filters=requested_filters,
                )

            return {
                "report": report_to_dict(report),
                "requested_url": url,
                "final_url": final_url,
                "html": html,
                "html_title": extract_html_title(html),
                "forms": form_field_summary(html),
                "filters": filter_status,
            }

    def _apply_report_filters(
        self,
        html: str,
        current_url: str,
        requested_filters: Dict[str, str],
    ) -> Tuple[str, str, Dict[str, Any]]:
        soup = BeautifulSoup(html, "html.parser")
        form = pick_report_form(soup)

        filter_status = {
            "requested": requested_filters,
            "applied": {},
            "not_applied": {},
            "submitted": False,
            "message": "",
        }

        if form is None:
            filter_status["not_applied"] = requested_filters
            filter_status["message"] = "El informe no ha devuelto ningún formulario donde aplicar filtros."
            return html, current_url, filter_status

        payload = collect_form_fields(form)
        fields = form_field_summary(str(form))[0]["fields"] if form_field_summary(str(form)) else []

        date_from = requested_filters.get("__date_from__")
        date_to = requested_filters.get("__date_to__")

        if date_from:
            field_name = find_matching_field(
                payload,
                fields,
                ("desde", "inicio", "inicial", "start", "from", "fecini", "fechainicio"),
            )
            if field_name:
                payload[field_name] = date_from
                filter_status["applied"]["date_from"] = {"field": field_name, "value": date_from}
            else:
                filter_status["not_applied"]["date_from"] = date_from

        if date_to:
            field_name = find_matching_field(
                payload,
                fields,
                ("hasta", "fin", "final", "end", "to", "fecfin", "fechafin"),
            )
            if field_name:
                payload[field_name] = date_to
                filter_status["applied"]["date_to"] = {"field": field_name, "value": date_to}
            else:
                filter_status["not_applied"]["date_to"] = date_to

        for key, value in requested_filters.items():
            if key in {"__date_from__", "__date_to__", "__month_from__", "__month_to__"}:
                continue

            if key in payload:
                payload[key] = value
                filter_status["applied"][key] = {"field": key, "value": value}
                continue

            normalized_key = normalize_key(key)
            matching_field = None
            for field_name in payload:
                if normalized_key and normalized_key in normalize_key(field_name):
                    matching_field = field_name
                    break

            if matching_field:
                payload[matching_field] = value
                filter_status["applied"][key] = {"field": matching_field, "value": value}
            else:
                filter_status["not_applied"][key] = value

        if not filter_status["applied"]:
            filter_status["message"] = (
                "Se recibieron filtros, pero no se encontró un campo compatible en el formulario. "
                "Usa /report-debug/{code} para ver los nombres exactos de los campos."
            )
            return html, current_url, filter_status

        submit_button = choose_submit_button(form)
        if submit_button:
            payload[submit_button[0]] = submit_button[1]

        action_url = urljoin(current_url, form.get("action") or current_url)
        method = clean_text(form.get("method") or "post").lower()
        headers = {"Referer": current_url}

        if method == "get":
            response = self.session.get(action_url, params=payload, headers=headers, timeout=REQUEST_TIMEOUT)
        else:
            response = self.session.post(action_url, data=payload, headers=headers, timeout=REQUEST_TIMEOUT)

        response.raise_for_status()
        filter_status["submitted"] = True
        filter_status["message"] = "Filtros enviados al formulario de Bifarma."
        return response.text, response.url, filter_status

    def _submit_pagination_event(
        self,
        html: str,
        current_url: str,
        event: Dict[str, Any],
    ) -> Tuple[str, str]:
        soup = BeautifulSoup(html, "html.parser")
        form = pick_report_form(soup)
        if form is None:
            raise BifarmaError("No se ha encontrado formulario ASP.NET para cambiar de página.")

        payload = collect_form_fields(form)
        payload["__EVENTTARGET"] = event["target"]
        payload["__EVENTARGUMENT"] = event["argument"]

        action_url = urljoin(current_url, form.get("action") or current_url)
        headers = {"Referer": current_url}
        response = self.session.post(action_url, data=payload, headers=headers, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.text, response.url

    def _submit_native_export_event(
        self,
        html: str,
        current_url: str,
        event: Dict[str, Any],
        expected_format: str,
    ) -> requests.Response:
        headers = {"Referer": current_url}

        if event["kind"] == "direct_url":
            export_url = urljoin(current_url, event["url"])
            response = self.session.get(export_url, headers=headers, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            response.raise_for_status()
            return response

        soup = BeautifulSoup(html, "html.parser")
        form = pick_report_form(soup)
        if form is None:
            raise BifarmaError("No se ha encontrado formulario ASP.NET para ejecutar la exportacion nativa.")

        payload = collect_form_fields(form)

        if event["kind"] == "postback":
            payload["__EVENTTARGET"] = event["target"]
            payload["__EVENTARGUMENT"] = event["argument"]
        elif event["kind"] == "form_button":
            payload[event["name"]] = event.get("value", "")
        else:
            raise BifarmaError(f"Tipo de evento de exportacion no soportado: {event['kind']}")

        action_url = urljoin(current_url, form.get("action") or current_url)
        response = self.session.post(action_url, data=payload, headers=headers, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response

    def export_report_native(
        self,
        code: str,
        requested_filters: Optional[Dict[str, str]],
        format_value: str,
    ) -> Dict[str, Any]:
        if format_value == "pdf":
            raise BifarmaError("La exportacion nativa solo esta disponible para xlsx/csv. Usa mode=html para PDF.")

        with self.lock:
            opened = self.open_report_html(code, requested_filters=requested_filters)
            events = find_native_export_events(opened["html"], preferred_format=format_value)

            if not events and format_value == "xlsx":
                events = find_native_export_events(opened["html"], preferred_format="csv")

            if not events:
                raise BifarmaError(
                    "No se ha encontrado boton/enlace de exportacion Excel o CSV en el HTML del informe. "
                    "Usa /report-debug/{code} para revisar los controles detectados."
                )

            attempts: List[Dict[str, Any]] = []
            for event in events[:8]:
                try:
                    response = self._submit_native_export_event(
                        html=opened["html"],
                        current_url=opened["final_url"],
                        event=event,
                        expected_format=format_value,
                    )
                except Exception as exc:
                    attempts.append({"event": event, "ok": False, "error": clean_text(exc)})
                    continue

                if response_looks_like_native_file(response, format_value):
                    content = response.content
                    return {
                        "opened": opened,
                        "content": content,
                        "media_type": native_media_type(format_value, content),
                        "filename": native_filename(opened["report"]["code"], format_value, content),
                        "event": event,
                        "attempts": attempts,
                        "content_type": response.headers.get("content-type", ""),
                        "content_disposition": response.headers.get("content-disposition", ""),
                    }

                attempts.append(
                    {
                        "event": event,
                        "ok": False,
                        "status_code": response.status_code,
                        "content_type": response.headers.get("content-type", ""),
                        "content_length": len(response.content or b""),
                    }
                )

            raise BifarmaError(
                "Se encontraron controles de exportacion, pero ninguno devolvio un archivo Excel/CSV valido. "
                f"Intentos: {json.dumps(attempts[:3], ensure_ascii=False)}"
            )

    def _browser_login(self, page: Any) -> None:
        username = self.username
        password = self.password
        if not username or not password:
            raise BifarmaError("Las credenciales de Bifarma son obligatorias.")

        page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60000)

        password_fields = page.locator("input[type='password']")
        if password_fields.count() < 1:
            raise BifarmaError("El navegador no ha encontrado el campo de password de Bifarma.")

        username_field = None
        username_candidates = page.locator("input[type='text'], input[type='email'], input:not([type])")
        for index in range(username_candidates.count()):
            candidate = username_candidates.nth(index)
            try:
                if candidate.is_visible(timeout=1000):
                    username_field = candidate
                    break
            except Exception:
                continue

        if username_field is None:
            raise BifarmaError("El navegador no ha encontrado el campo de usuario de Bifarma.")

        username_field.fill(username)
        password_field = password_fields.first
        password_field.fill(password)

        clicked = False
        submit_selectors = [
            "input[type='submit']",
            "button[type='submit']",
            "input[value*='Entrar']",
            "input[value*='Acceder']",
            "button:has-text('Entrar')",
            "button:has-text('Acceder')",
        ]
        for selector in submit_selectors:
            locator = page.locator(selector)
            if locator.count() < 1:
                continue
            try:
                first = locator.first
                if first.is_visible(timeout=1000):
                    first.click(timeout=10000)
                    clicked = True
                    break
            except Exception:
                continue

        if not clicked:
            password_field.press("Enter")

        try:
            page.wait_for_load_state("networkidle", timeout=60000)
        except Exception:
            pass

        if "loginform.aspx" in page.url.lower():
            raise BifarmaError("Bifarma sigue mostrando la pantalla de login. Revisa las credenciales.")

    def _browser_fill_filter(self, page: Any, hints: List[str], value: str) -> Optional[str]:
        if not value:
            return None

        return page.evaluate(
            """
            ({ hints, value }) => {
                const normalize = (text) => (text || "")
                    .toLowerCase()
                    .normalize("NFD")
                    .replace(/[\\u0300-\\u036f]/g, "")
                    .replace(/[^a-z0-9]+/g, "");

                const isVisible = (el) => {
                    const style = window.getComputedStyle(el);
                    const rect = el.getBoundingClientRect();
                    return style.visibility !== "hidden"
                        && style.display !== "none"
                        && rect.width > 0
                        && rect.height > 0;
                };

                const labelText = (el) => {
                    let text = "";
                    if (el.id) {
                        const label = document.querySelector(`label[for="${CSS.escape(el.id)}"]`);
                        if (label) text += " " + label.innerText;
                    }
                    const parentText = el.closest("td, div, span, label")?.innerText || "";
                    return `${text} ${parentText}`;
                };

                const controls = Array.from(document.querySelectorAll("input:not([type='hidden']), textarea"));
                let best = null;
                let bestScore = 0;

                for (const el of controls) {
                    if (!isVisible(el)) continue;
                    const haystack = normalize([
                        el.name,
                        el.id,
                        el.placeholder,
                        el.title,
                        el.getAttribute("aria-label"),
                        labelText(el),
                    ].join(" "));

                    let score = 0;
                    for (const hint of hints) {
                        if (haystack.includes(normalize(hint))) score += 10;
                    }
                    if (haystack.includes("fecha") || haystack.includes("date") || haystack.includes("fec")) {
                        score += 3;
                    }
                    if (score > bestScore) {
                        best = el;
                        bestScore = score;
                    }
                }

                if (!best) return null;
                best.focus();
                best.value = value;
                best.dispatchEvent(new Event("input", { bubbles: true }));
                best.dispatchEvent(new Event("change", { bubbles: true }));
                best.blur();
                return best.name || best.id || best.placeholder || "campo";
            }
            """,
            {"hints": hints, "value": value},
        )

    # JS: descubre globales DevExpress con items YYYYMM (combos Ano-Mes).
    _DISCOVER_YYYYMM_COMBOS_JS = r"""
    () => {
      var out = [];
      for (var k in window) {
        var v = window[k];
        if (!v || typeof v !== 'object') continue;
        if (typeof v.SetValue !== 'function' || typeof v.GetText !== 'function') continue;
        if (typeof v.GetItemCount !== 'function' || v.GetItemCount() < 2) continue;
        var sample = [];
        for (var i = 0; i < Math.min(v.GetItemCount(), 3); i++) {
          try { sample.push(String(v.GetItem(i).value)); } catch (e) {}
        }
        if (!sample.every(function(s) { return /^20\d{4}$/.test(s); })) continue;
        var el = (typeof v.GetMainElement === 'function') ? v.GetMainElement() : null;
        out.push({
          key: k,
          mainId: el ? el.id : null,
          nItems: v.GetItemCount(),
          value: String(v.GetValue()),
          text: v.GetText()
        });
      }
      return out;
    }
    """

    # JS: lista todos los items de una combo YYYYMM.
    _LIST_YYYYMM_ITEMS_JS = r"""
    (key) => {
      var c = window[key];
      if (!c) return [];
      var items = [];
      for (var i = 0; i < c.GetItemCount(); i++) {
        var it = c.GetItem(i);
        items.push({ v: String(it.value), t: it.text });
      }
      return items;
    }
    """

    def _browser_select_yyyymm_combo(self, page: Any, key: str, main_id: str, value: str, text: str) -> bool:
        """Abre la combo DevExpress con ShowDropDown() y hace clic real en el item."""
        try:
            page.evaluate(f"() => window['{key}'].ShowDropDown()")
            page.wait_for_timeout(500)
            listbox = f"#{main_id}_DDD_L"
            try:
                page.wait_for_selector(listbox, state="visible", timeout=5000)
            except Exception:
                return False
            item = page.locator(f"{listbox} td.dxeListBoxItem_Glass").filter(
                has_text=re.compile(rf"^{re.escape(text)}$")
            )
            item.first.click(timeout=8000)
            page.wait_for_timeout(1200)
            return True
        except Exception:
            return False

    # JS: descubre date-edit DevExpress vivos (tienen SetDate/GetDate). UN.080
    # usa dos ASPxDateEdit (_biCt4=Desde, _biCt5=Hasta), no las combos de UN.088.
    _DISCOVER_DATEEDITS_JS = r"""
    () => {
      var out = [];
      for (var k in window) {
        var v;
        try { v = window[k]; } catch (e) { continue; }
        if (!v || typeof v !== 'object') continue;
        if (typeof v.SetDate !== 'function' || typeof v.GetDate !== 'function') continue;
        var el = (typeof v.GetMainElement === 'function') ? v.GetMainElement() : null;
        out.push({ key: k, mainId: el ? el.id : null });
      }
      return out;
    }
    """

    # JS: fija la fecha de un date-edit por su key global. SetDate commitea el
    # input visible (verificado en spike): el postback lo serializa correcto.
    _SET_DATEEDIT_JS = r"""
    ({ key, y, m, d }) => {
      var c = window[key];
      if (!c || typeof c.SetDate !== 'function') return false;
      try { c.SetDate(new Date(y, m - 1, d)); return true; } catch (e) { return false; }
    }
    """

    @staticmethod
    def _control_suffix(key: str) -> int:
        nums = re.findall(r"\d+", key)
        return int(nums[-1]) if nums else 0

    def _browser_set_dateedit(self, page: Any, key: str, value_ddmmyyyy: str) -> bool:
        parsed = parse_ddmmyyyy(value_ddmmyyyy)
        if not parsed:
            return False
        year, month, day = parsed
        return bool(page.evaluate(self._SET_DATEEDIT_JS, {"key": key, "y": year, "m": month, "d": day}))

    def _browser_apply_filters(self, page: Any, requested_filters: Optional[Dict[str, str]]) -> Dict[str, Any]:
        status = {
            "requested": requested_filters or {},
            "applied": {},
            "not_applied": {},
        }
        if not requested_filters:
            return status

        month_from = requested_filters.get("__month_from__")
        month_to = requested_filters.get("__month_to__")

        if month_from or month_to:
            combos = page.evaluate(self._DISCOVER_YYYYMM_COMBOS_JS)
            if len(combos) >= 2:
                def _suffix(key: str) -> int:
                    nums = re.findall(r"\d+", key)
                    return int(nums[-1]) if nums else 0
                combos.sort(key=lambda c: _suffix(c["key"]))
                desde_combo, hasta_combo = combos[0], combos[1]

                if month_from:
                    items = page.evaluate(self._LIST_YYYYMM_ITEMS_JS, desde_combo["key"])
                    match = next((it for it in items if it["v"] == month_from), None)
                    if match and self._browser_select_yyyymm_combo(
                        page, desde_combo["key"], desde_combo["mainId"], match["v"], match["t"]
                    ):
                        status["applied"]["month_from"] = {"field": desde_combo["key"], "value": month_from, "text": match["t"]}
                    else:
                        status["not_applied"]["month_from"] = month_from

                if month_to:
                    items = page.evaluate(self._LIST_YYYYMM_ITEMS_JS, hasta_combo["key"])
                    match = next((it for it in items if it["v"] == month_to), None)
                    if match and self._browser_select_yyyymm_combo(
                        page, hasta_combo["key"], hasta_combo["mainId"], match["v"], match["t"]
                    ):
                        status["applied"]["month_to"] = {"field": hasta_combo["key"], "value": month_to, "text": match["t"]}
                    else:
                        status["not_applied"]["month_to"] = month_to
            else:
                if month_from:
                    status["not_applied"]["month_from"] = month_from
                if month_to:
                    status["not_applied"]["month_to"] = month_to

        date_from = requested_filters.get("__date_from__")
        date_to = requested_filters.get("__date_to__")

        if date_from or date_to:
            # Ruta DevExpress (UN.080): dos ASPxDateEdit; el primero por sufijo de
            # control es Desde y el segundo Hasta. SetValue seria silencioso, como
            # en las combos de UN.088; SetDate si commitea el input del postback.
            edits = page.evaluate(self._DISCOVER_DATEEDITS_JS)
            edits.sort(key=lambda e: self._control_suffix(e["key"]))

            if edits:
                if date_from:
                    if self._browser_set_dateedit(page, edits[0]["key"], date_from):
                        status["applied"]["date_from"] = {"field": edits[0]["key"], "value": date_from}
                    else:
                        status["not_applied"]["date_from"] = date_from
                if date_to:
                    if len(edits) >= 2 and self._browser_set_dateedit(page, edits[1]["key"], date_to):
                        status["applied"]["date_to"] = {"field": edits[1]["key"], "value": date_to}
                    else:
                        status["not_applied"]["date_to"] = date_to
            else:
                # Fallback heuristico para formularios no-DevExpress.
                if date_from:
                    field = self._browser_fill_filter(
                        page,
                        ["fecha desde", "desde", "inicio", "inicial", "date from", "start", "fecini"],
                        date_from,
                    )
                    if field:
                        status["applied"]["date_from"] = {"field": field, "value": date_from}
                    else:
                        status["not_applied"]["date_from"] = date_from
                if date_to:
                    field = self._browser_fill_filter(
                        page,
                        ["fecha hasta", "hasta", "fin", "final", "date to", "end", "fecfin"],
                        date_to,
                    )
                    if field:
                        status["applied"]["date_to"] = {"field": field, "value": date_to}
                    else:
                        status["not_applied"]["date_to"] = date_to

        for key, value in requested_filters.items():
            if key in {"__date_from__", "__date_to__", "__month_from__", "__month_to__"}:
                continue
            field = self._browser_fill_filter(page, [key], value)
            if field:
                status["applied"][key] = {"field": field, "value": value}
            else:
                status["not_applied"][key] = value

        return status

    def _browser_click_update_data(self, page: Any) -> None:
        patterns = [
            r"^Actualizar datos$",
            r"^Actualizar$",
            r"^Consultar$",
            r"^Buscar$",
            r"^Aceptar$",
        ]
        for pattern in patterns:
            locator = page.get_by_text(re.compile(pattern, re.IGNORECASE))
            for index in range(min(locator.count(), 4)):
                item = locator.nth(index)
                try:
                    if item.is_visible(timeout=1000):
                        item.click(timeout=10000)
                        try:
                            page.wait_for_load_state("networkidle", timeout=45000)
                        except Exception:
                            pass
                        page.wait_for_timeout(1000)
                        return
                except Exception:
                    continue

        # UN.080: el boton "Actualizar datos" es un ASPxButton cuyo <input submit>
        # esta oculto para Playwright, y dispara un postback de pagina COMPLETA
        # (no un callback XHR). El clic por texto de arriba no lo alcanza; aqui lo
        # disparamos via __doPostBack sobre su name y esperamos la navegacion.
        self._browser_postback_update(page)

    def _browser_postback_update(self, page: Any) -> None:
        name = page.evaluate(
            """
            () => {
                const inp = document.querySelector(
                    'input[type=submit][value*="Actualizar"], input[type=submit][value*="Consultar"]'
                );
                return inp ? inp.name : null;
            }
            """
        )
        if not name:
            return
        try:
            with page.expect_navigation(wait_until="load", timeout=45000):
                # setTimeout: que evaluate retorne antes de que el postback navegue.
                page.evaluate(
                    "(n) => { setTimeout(function(){ window.__doPostBack(n, ''); }, 0); }",
                    name,
                )
        except Exception:
            pass
        try:
            page.wait_for_load_state("networkidle", timeout=30000)
        except Exception:
            pass
        page.wait_for_timeout(1500)

    def _browser_try_download_from_visible_menu(self, page: Any, format_value: str) -> Optional[Any]:
        if format_value == "csv":
            patterns = [r"\bCSV\b", r"Exportar.*CSV", r"Descargar.*CSV"]
        else:
            patterns = [r"\bXLSX\b", r"\bXLS\b", r"Excel", r"Exportar.*Excel", r"Descargar.*Excel"]

        for pattern in patterns:
            locator = page.get_by_text(re.compile(pattern, re.IGNORECASE))
            for index in range(min(locator.count(), 8)):
                item = locator.nth(index)
                try:
                    if not item.is_visible(timeout=1000):
                        continue
                    with page.expect_download(timeout=30000) as download_info:
                        item.click(timeout=10000)
                    return download_info.value
                except Exception:
                    continue

        return None

    def _browser_visible_export_texts(self, page: Any, limit: int = 120) -> List[str]:
        try:
            values = page.evaluate(
                """
                ({ limit }) => {
                    const isVisible = (el) => {
                        const style = window.getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        return style.visibility !== "hidden"
                            && style.display !== "none"
                            && rect.width > 0
                            && rect.height > 0;
                    };
                    const interesting = /(excel|xlsx|xls|csv|pdf|export|exportar|descargar|download|guardar)/i;
                    const result = [];
                    for (const el of Array.from(document.querySelectorAll("a,button,input,div,span,td,li"))) {
                        if (!isVisible(el)) continue;
                        const text = [
                            el.innerText,
                            el.value,
                            el.title,
                            el.alt,
                            el.getAttribute("aria-label"),
                            el.id,
                            el.name,
                            el.className,
                        ].filter(Boolean).join(" ").replace(/\\s+/g, " ").trim();
                        if (text && interesting.test(text)) {
                            result.push(text.slice(0, 220));
                        }
                        if (result.length >= limit) break;
                    }
                    return Array.from(new Set(result));
                }
                """,
                {"limit": limit},
            )
            return [clean_text(value) for value in values if clean_text(value)]
        except Exception as exc:
            return [f"No se pudo leer texto visible: {clean_text(exc)}"]

    def _browser_context_menu_debug(self, page: Any, format_value: str) -> Dict[str, Any]:
        grid_selectors = [
            ".dxgvDataRow td",
            ".dxgvAltDataRow td",
            "[class*='dxgvDataRow'] td",
            "[class*='dxgvAltDataRow'] td",
            ".dxgvDataRow",
            ".dxgvTable td",
            ".dxgvTable",
            ".dxgvControl",
            "[id*='Grid'] td",
            "[id*='grid'] td",
            "[id*='Grid']",
            "[id*='grid']",
            "table td",
            "table",
        ]
        attempts: List[Dict[str, Any]] = []

        for selector in grid_selectors:
            locator = page.locator(selector)
            count = locator.count()
            if count < 1:
                attempts.append({"selector": selector, "count": 0, "clicked": False})
                continue

            for index in range(min(count, 3)):
                candidate = locator.nth(index)
                attempt: Dict[str, Any] = {"selector": selector, "index": index, "count": count, "clicked": False}
                try:
                    if not candidate.is_visible(timeout=1000):
                        attempt["visible"] = False
                        attempts.append(attempt)
                        continue

                    box = candidate.bounding_box()
                    attempt["visible"] = True
                    attempt["box"] = box
                    candidate.click(button="right", timeout=10000)
                    page.wait_for_timeout(900)
                    attempt["clicked"] = True
                    attempt["visible_export_texts"] = self._browser_visible_export_texts(page, limit=60)

                    export_menu_count = page.get_by_text(
                        re.compile(r"Exportar|Export|Descargar|Download|Excel|XLS|XLSX|CSV", re.IGNORECASE)
                    ).count()
                    attempt["export_text_locator_count"] = export_menu_count
                    attempts.append(attempt)

                    if attempt["visible_export_texts"]:
                        return {
                            "ok": True,
                            "format": format_value,
                            "page_url": page.url,
                            "page_title": page.title(),
                            "message": "Se ha abierto el menu contextual y se han visto textos de exportacion.",
                            "attempts": attempts,
                            "visible_export_texts": attempt["visible_export_texts"],
                        }
                except Exception as exc:
                    attempt["error"] = clean_text(exc)
                    attempts.append(attempt)

        return {
            "ok": False,
            "format": format_value,
            "page_url": page.url,
            "page_title": page.title(),
            "message": "No se han visto opciones de exportacion despues del clic derecho.",
            "attempts": attempts[-20:],
            "visible_export_texts": self._browser_visible_export_texts(page, limit=80),
        }

    def _browser_open_context_menu_and_download(self, page: Any, format_value: str) -> Any:
        grid_selectors = [
            ".dxgvDataRow td",
            ".dxgvAltDataRow td",
            "[class*='dxgvDataRow'] td",
            "[class*='dxgvAltDataRow'] td",
            ".dxgvDataRow",
            ".dxgvAltDataRow",
            ".dxgvTable td",
            ".dxgvTable",
            ".dxgvControl",
            "[id*='Grid'] td",
            "[id*='grid'] td",
            "[id*='Grid']",
            "[id*='grid']",
            "table td",
            "table",
        ]

        last_error = ""
        for selector in grid_selectors:
            locator = page.locator(selector)
            if locator.count() < 1:
                continue

            for index in range(min(locator.count(), 5)):
                candidate = locator.nth(index)
                try:
                    if not candidate.is_visible(timeout=1000):
                        continue
                    candidate.click(button="right", timeout=10000)
                    page.wait_for_timeout(750)

                    download = self._browser_try_download_from_visible_menu(page, format_value)
                    if download:
                        return download

                    export_menu = page.get_by_text(re.compile(r"Exportar|Export|Descargar|Download", re.IGNORECASE))
                    for menu_index in range(min(export_menu.count(), 6)):
                        menu_item = export_menu.nth(menu_index)
                        try:
                            if menu_item.is_visible(timeout=1000):
                                menu_item.click(timeout=10000)
                                page.wait_for_timeout(750)
                                download = self._browser_try_download_from_visible_menu(page, format_value)
                                if download:
                                    return download
                        except Exception as exc:
                            last_error = clean_text(exc)
                            continue
                except Exception as exc:
                    last_error = clean_text(exc)
                    continue

        raise BifarmaError(
            "El navegador ha abierto el informe, pero no ha conseguido descargar desde el menu contextual. "
            f"Ultimo detalle: {last_error}. "
            f"Textos visibles relacionados: {self._browser_visible_export_texts(page, limit=30)}"
        )

    def export_report_browser(
        self,
        code: str,
        requested_filters: Optional[Dict[str, str]],
        format_value: str,
    ) -> Dict[str, Any]:
        if format_value == "pdf":
            raise BifarmaError("browser_export solo esta disponible para xlsx/csv.")

        try:
            from playwright.sync_api import sync_playwright
        except Exception as exc:
            raise BifarmaError(
                "El modo browser_export necesita Playwright. En Render instala la dependencia y usa "
                "el Build Command: bash render-build.sh"
            ) from exc

        chromium_status = chromium_launch_status(auto_install=True)
        if not chromium_status.get("ok"):
            raise BifarmaError(
                "Playwright esta instalado, pero Chromium no arranca en el servidor. "
                f"Detalle: {chromium_status.get('error')}. "
                "En Render usa Build Command: bash render-build.sh y vuelve a desplegar."
            )

        report = self.find_report(code)
        report_url = make_document_url(report.role, report.code)

        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            context = browser.new_context(accept_downloads=True, viewport={"width": 1440, "height": 1000})
            page = context.new_page()
            try:
                self._browser_login(page)
                page.goto(report_url, wait_until="domcontentloaded", timeout=60000)
                try:
                    page.wait_for_load_state("networkidle", timeout=60000)
                except Exception:
                    pass
                page.wait_for_timeout(1500)

                filter_status = self._browser_apply_filters(page, requested_filters)
                if filter_status["applied"]:
                    self._browser_click_update_data(page)

                download = self._browser_open_context_menu_and_download(page, format_value)
                download_path = download.path()
                content = Path(download_path).read_bytes()

                if not content:
                    raise BifarmaError("Bifarma ha generado una descarga vacia.")

                return {
                    "content": content,
                    "media_type": native_media_type(format_value, content),
                    "filename": native_filename(report.code, format_value, content),
                    "report": report_to_dict(report),
                    "source_url": page.url,
                    "filters": filter_status,
                }
            finally:
                context.close()
                browser.close()

    def debug_report_browser_export(
        self,
        code: str,
        requested_filters: Optional[Dict[str, str]],
        format_value: str,
    ) -> Dict[str, Any]:
        if format_value == "pdf":
            raise BifarmaError("browser_export_debug solo esta disponible para xlsx/csv.")

        try:
            from playwright.sync_api import sync_playwright
        except Exception as exc:
            raise BifarmaError(
                "El diagnostico por navegador necesita Playwright. En Render usa Build Command: bash render-build.sh"
            ) from exc

        chromium_status = chromium_launch_status(auto_install=True)
        if not chromium_status.get("ok"):
            raise BifarmaError(
                "Playwright esta instalado, pero Chromium no arranca en el servidor. "
                f"Detalle: {chromium_status.get('error')}"
            )

        report = self.find_report(code)
        report_url = make_document_url(report.role, report.code)
        steps: List[Dict[str, Any]] = []

        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            context = browser.new_context(accept_downloads=True, viewport={"width": 1440, "height": 1000})
            page = context.new_page()
            try:
                self._browser_login(page)
                steps.append({"step": "login", "ok": True, "url": page.url})

                page.goto(report_url, wait_until="domcontentloaded", timeout=60000)
                try:
                    page.wait_for_load_state("networkidle", timeout=60000)
                except Exception as exc:
                    steps.append({"step": "networkidle", "ok": False, "error": clean_text(exc)})
                page.wait_for_timeout(1500)
                steps.append({"step": "open_report", "ok": True, "url": page.url, "title": page.title()})

                filter_status = self._browser_apply_filters(page, requested_filters)
                steps.append({"step": "filters", "ok": True, "filters": filter_status})
                if filter_status["applied"]:
                    self._browser_click_update_data(page)
                    steps.append({"step": "update_data", "ok": True, "url": page.url})

                context_debug = self._browser_context_menu_debug(page, format_value)
                steps.append({"step": "context_menu", "ok": context_debug.get("ok"), "message": context_debug.get("message")})

                return {
                    "ok": context_debug.get("ok", False),
                    "version": APP_VERSION,
                    "report": report_to_dict(report),
                    "format": format_value,
                    "steps": steps,
                    "context_menu": context_debug,
                    "chromium": chromium_status,
                }
            finally:
                context.close()
                browser.close()

    def collect_report_data(
        self,
        code: str,
        requested_filters: Optional[Dict[str, str]],
        max_rows: int,
        all_pages: bool,
        max_pages: int,
    ) -> Dict[str, Any]:
        with self.lock:
            opened = self.open_report_html(code, requested_filters=requested_filters)
            html = opened["html"]
            current_url = opened["final_url"]
            columns: List[str] = []
            rows: List[Dict[str, Any]] = []
            page_details: List[Dict[str, Any]] = []
            seen_signatures = set()
            current_page = 1
            stopped_reason = "single_page_requested"
            pagination_hint = extract_pagination_hint(html)

            while True:
                remaining_rows = max_rows - len(rows)
                if remaining_rows <= 0:
                    stopped_reason = "max_rows_reached"
                    break

                extracted = extract_tables(html, max_rows=remaining_rows)
                page_rows = extracted["rows"]

                if current_page > 1 and extracted["table_index"] is None:
                    stopped_reason = "next_page_without_table"
                    break

                signature = rows_signature(page_rows)
                if current_page > 1 and signature and signature in seen_signatures:
                    stopped_reason = "repeated_page_detected"
                    break

                if signature:
                    seen_signatures.add(signature)

                columns = merge_columns(columns, extracted["columns"])
                normalized_page_rows = normalize_rows_to_columns(page_rows, columns)
                rows.extend(normalized_page_rows[:remaining_rows])

                page_details.append(
                    {
                        "page": current_page,
                        "table_index": extracted["table_index"],
                        "tables_detected": extracted["tables_detected"],
                        "rows_read": len(normalized_page_rows[:remaining_rows]),
                        "source_url": current_url,
                    }
                )

                latest_hint = extract_pagination_hint(html)
                for key, value in latest_hint.items():
                    if value:
                        pagination_hint[key] = value

                if not all_pages:
                    stopped_reason = "single_page_requested"
                    break

                if len(rows) >= max_rows:
                    stopped_reason = "max_rows_reached"
                    break

                if current_page >= max_pages:
                    stopped_reason = "max_pages_reached"
                    break

                next_event = find_next_pagination_event(html, current_page)
                if not next_event:
                    stopped_reason = "no_next_page_event_found"
                    break

                html, current_url = self._submit_pagination_event(html, current_url, next_event)
                current_page += 1

            return {
                "opened": opened,
                "extracted": {
                    "table_index": page_details[0]["table_index"] if page_details else None,
                    "tables_detected": page_details[0]["tables_detected"] if page_details else 0,
                    "columns": columns,
                    "rows": rows[:max_rows],
                    "total_rows": len(rows),
                },
                "pagination": {
                    "enabled": all_pages,
                    "pages_read": len(page_details),
                    "max_pages": max_pages,
                    "stopped_reason": stopped_reason,
                    **pagination_hint,
                    "pages": page_details,
                },
            }


def get_client(
    x_bifarma_user: Optional[str] = Header(default=None, alias="x-bifarma-user"),
    x_bifarma_password: Optional[str] = Header(default=None, alias="x-bifarma-password"),
) -> BifarmaClient:
    if not x_bifarma_user or not x_bifarma_password:
        raise HTTPException(status_code=401, detail="Headers x-bifarma-user y x-bifarma-password son obligatorios.")
    return BifarmaClient(x_bifarma_user, x_bifarma_password)


def handle_bifarma_error(exc: Exception) -> JSONResponse:
    if isinstance(exc, HTTPException):
        raise exc
    return JSONResponse(status_code=502, content={"ok": False, "error": clean_text(exc)})


def generate_export_file(
    client: BifarmaClient,
    code: str,
    requested_filters: Dict[str, str],
    format: str,
    mode: str,
    max_rows: int,
    all_pages: bool,
    max_pages: int,
) -> Dict[str, Any]:
    if mode in {"auto", "native_export"} and format in {"xlsx", "csv"}:
        try:
            native = client.export_report_native(
                code=code,
                requested_filters=requested_filters,
                format_value=format,
            )
            return {
                "content": native["content"],
                "media_type": native["media_type"],
                "filename": native["filename"],
                "export_mode": "native_export",
                "report": native["opened"]["report"],
            }
        except Exception as native_exc:
            if mode == "native_export":
                try:
                    browser_export = client.export_report_browser(
                        code=code,
                        requested_filters=requested_filters,
                        format_value=format,
                    )
                    return {
                        "content": browser_export["content"],
                        "media_type": browser_export["media_type"],
                        "filename": browser_export["filename"],
                        "export_mode": "browser_export_after_native",
                        "report": browser_export["report"],
                    }
                except Exception as browser_exc:
                    raise BifarmaError(
                        "No se ha podido exportar por boton HTML ni por menu contextual. "
                        f"Error exportacion HTML: {clean_text(native_exc)}. "
                        f"Error navegador: {clean_text(browser_exc)}"
                    ) from browser_exc

    if mode in {"auto", "browser_export"} and format in {"xlsx", "csv"}:
        try:
            browser_export = client.export_report_browser(
                code=code,
                requested_filters=requested_filters,
                format_value=format,
            )
            return {
                "content": browser_export["content"],
                "media_type": browser_export["media_type"],
                "filename": browser_export["filename"],
                "export_mode": "browser_export",
                "report": browser_export["report"],
            }
        except Exception:
            if mode == "browser_export":
                raise

    collected = client.collect_report_data(
        code=code,
        requested_filters=requested_filters,
        max_rows=max_rows,
        all_pages=all_pages,
        max_pages=max_pages,
    )
    opened = collected["opened"]
    extracted = collected["extracted"]
    columns = extracted["columns"]
    rows = extracted["rows"]
    filename_base = safe_filename(opened["report"]["code"])

    if format == "csv":
        content = build_csv_bytes(columns, rows)
        media_type = "text/csv; charset=utf-8"
        filename = f"{filename_base}.csv"
    elif format == "pdf":
        content = build_pdf_bytes(opened["report"]["title"], columns, rows)
        media_type = "application/pdf"
        filename = f"{filename_base}.pdf"
    else:
        content = build_xlsx_bytes(columns, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{filename_base}.xlsx"

    return {
        "content": content,
        "media_type": media_type,
        "filename": filename,
        "export_mode": "html",
        "report": opened["report"],
        "pagination": collected["pagination"],
    }


@app.get("/")
def root() -> Dict[str, Any]:
    return {
        "ok": True,
        "name": "Bifarma Connector API",
        "version": APP_VERSION,
        "docs": "/docs",
        "status": "/status",
    }


@app.get("/health")
def health() -> Dict[str, Any]:
    """Endpoint de salud sin autenticacion. Permite detectar si el servicio
    esta listo tras el cold start de Render free tier sin exponer la API key."""
    return {"ok": True, "version": APP_VERSION}


@app.get("/status", dependencies=[Depends(require_api_key)])
def status() -> Dict[str, Any]:
    return {
        "ok": True,
        "version": APP_VERSION,
        "base_url": BASE_URL,
        "api_key_configured": bool(os.getenv("API_KEY")),
        "stateless": True,
        "playwright": playwright_import_status(),
    }


@app.get("/browser-status", dependencies=[Depends(require_api_key)])
def browser_status(
    auto_install: bool = Query(False, description="Si es true, intenta instalar Chromium si falta."),
) -> Dict[str, Any]:
    launch = chromium_launch_status(auto_install=auto_install)
    return {
        "ok": launch["ok"],
        "version": APP_VERSION,
        "playwright": playwright_import_status(),
        "chromium": launch,
    }


@app.post("/login", dependencies=[Depends(require_api_key)])
def login(client: BifarmaClient = Depends(get_client)) -> JSONResponse:
    try:
        result = client.ensure_login(force=True)
        return JSONResponse({"ok": True, **result})
    except Exception as exc:
        return handle_bifarma_error(exc)


@app.get("/reports-menu", dependencies=[Depends(require_api_key)])
def reports_menu(
    refresh: bool = Query(False, description="Fuerza volver a leer el menú privado de Bifarma."),
    client: BifarmaClient = Depends(get_client),
) -> JSONResponse:
    try:
        return JSONResponse({"ok": True, **client.reports_menu(force_refresh=refresh)})
    except Exception as exc:
        return handle_bifarma_error(exc)


@app.get("/report-debug/{code}", dependencies=[Depends(require_api_key)])
def report_debug(code: str, client: BifarmaClient = Depends(get_client)) -> JSONResponse:
    try:
        opened = client.open_report_html(code)
        pagination_events = [
            event for event in find_postback_events(opened["html"]) if event["is_pagination"]
        ][:30]
        native_export_events = find_native_export_events(opened["html"], preferred_format="xlsx")[:30]
        return JSONResponse(
            {
                "ok": True,
                "report": opened["report"],
                "requested_url": opened["requested_url"],
                "final_url": opened["final_url"],
                "html_title": opened["html_title"],
                "forms": opened["forms"],
                "pagination": {
                    "hint": extract_pagination_hint(opened["html"]),
                    "events_detected": pagination_events,
                },
                "native_export": {
                    "events_detected": native_export_events,
                    "script_snippets": extract_export_script_snippets(opened["html"]),
                },
            }
        )
    except Exception as exc:
        return handle_bifarma_error(exc)


@app.get("/browser-export-debug/{code}", dependencies=[Depends(require_api_key)])
def browser_export_debug(
    code: str,
    request: Request,
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    start_date: Optional[str] = Query(None, description="Fecha inicial. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    end_date: Optional[str] = Query(None, description="Fecha final. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    date_from: Optional[str] = Query(None, description="Alias de start_date."),
    date_to: Optional[str] = Query(None, description="Alias de end_date."),
    filters: Optional[str] = Query(None, description="JSON con filtros explicitos por nombre de campo."),
    client: BifarmaClient = Depends(get_client),
) -> JSONResponse:
    try:
        requested_filters = build_requested_filters(request, start_date, end_date, date_from, date_to, filters)
        result = client.debug_report_browser_export(
            code=code,
            requested_filters=requested_filters,
            format_value=format,
        )
        return JSONResponse(result)
    except Exception as exc:
        return handle_bifarma_error(exc)


@app.get("/report-open/{code}", dependencies=[Depends(require_api_key)])
def report_open(
    code: str,
    request: Request,
    start_date: Optional[str] = Query(None, description="Fecha inicial. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    end_date: Optional[str] = Query(None, description="Fecha final. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    date_from: Optional[str] = Query(None, description="Alias de start_date."),
    date_to: Optional[str] = Query(None, description="Alias de end_date."),
    filters: Optional[str] = Query(None, description="JSON con filtros explícitos por nombre de campo."),
    client: BifarmaClient = Depends(get_client),
) -> JSONResponse:
    try:
        requested_filters = build_requested_filters(request, start_date, end_date, date_from, date_to, filters)
        opened = client.open_report_html(code, requested_filters=requested_filters)
        return JSONResponse(
            {
                "ok": True,
                "report": opened["report"],
                "requested_url": opened["requested_url"],
                "final_url": opened["final_url"],
                "html_title": opened["html_title"],
                "filters": opened["filters"],
                "html_length": len(opened["html"]),
                "forms_detected": len(opened["forms"]),
            }
        )
    except Exception as exc:
        return handle_bifarma_error(exc)


@app.get("/report-data/{code}", dependencies=[Depends(require_api_key)])
def report_data(
    code: str,
    request: Request,
    max_rows: int = Query(50, ge=1, le=50000),
    all_pages: bool = Query(False, description="Si es true, intenta recorrer todas las páginas del informe."),
    max_pages: int = Query(100, ge=1, le=2000, description="Límite de páginas a recorrer cuando all_pages=true."),
    start_date: Optional[str] = Query(None, description="Fecha inicial. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    end_date: Optional[str] = Query(None, description="Fecha final. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    date_from: Optional[str] = Query(None, description="Alias de start_date."),
    date_to: Optional[str] = Query(None, description="Alias de end_date."),
    filters: Optional[str] = Query(None, description="JSON con filtros explícitos por nombre de campo."),
    client: BifarmaClient = Depends(get_client),
) -> JSONResponse:
    try:
        requested_filters = build_requested_filters(request, start_date, end_date, date_from, date_to, filters)
        collected = client.collect_report_data(
            code=code,
            requested_filters=requested_filters,
            max_rows=max_rows,
            all_pages=all_pages,
            max_pages=max_pages,
        )
        opened = collected["opened"]
        extracted = collected["extracted"]

        return JSONResponse(
            {
                "ok": True,
                "report": opened["report"],
                "source_url": opened["final_url"],
                "html_title": opened["html_title"],
                "filters": opened["filters"],
                "pagination": collected["pagination"],
                "table_index": extracted["table_index"],
                "tables_detected": extracted["tables_detected"],
                "columns": extracted["columns"],
                "total_rows": extracted["total_rows"],
                "returned_rows": len(extracted["rows"]),
                "rows": extracted["rows"],
            }
        )
    except Exception as exc:
        return handle_bifarma_error(exc)


@app.get("/report-export/{code}", dependencies=[Depends(require_api_key)])
def report_export(
    code: str,
    request: Request,
    format: str = Query("xlsx", pattern="^(csv|xlsx|pdf)$"),
    mode: str = Query(
        "auto",
        pattern="^(auto|native_export|browser_export|html)$",
        description="auto intenta primero exportacion nativa, despues navegador con menu contextual y usa HTML como respaldo; native_export intenta Excel/CSV nativo y, si no hay boton HTML, cae a browser_export; browser_export usa navegador y clic derecho; html usa la extraccion anterior.",
    ),
    max_rows: int = Query(500, ge=1, le=50000),
    all_pages: bool = Query(False, description="Si es true, intenta recorrer todas las páginas del informe."),
    max_pages: int = Query(100, ge=1, le=2000, description="Límite de páginas a recorrer cuando all_pages=true."),
    start_date: Optional[str] = Query(None, description="Fecha inicial. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    end_date: Optional[str] = Query(None, description="Fecha final. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    date_from: Optional[str] = Query(None, description="Alias de start_date."),
    date_to: Optional[str] = Query(None, description="Alias de end_date."),
    filters: Optional[str] = Query(None, description="JSON con filtros explícitos por nombre de campo."),
    client: BifarmaClient = Depends(get_client),
) -> StreamingResponse:
    try:
        requested_filters = build_requested_filters(request, start_date, end_date, date_from, date_to, filters)
        exported = generate_export_file(
            client=client,
            code=code,
            requested_filters=requested_filters,
            format=format,
            mode=mode,
            max_rows=max_rows,
            all_pages=all_pages,
            max_pages=max_pages,
        )
        return StreamingResponse(
            io.BytesIO(exported["content"]),
            media_type=exported["media_type"],
            headers={
                "Content-Disposition": f'attachment; filename="{exported["filename"]}"',
                "X-Bifarma-Export-Mode": exported["export_mode"],
            },
        )
    except Exception as exc:
        if isinstance(exc, HTTPException):
            raise exc
        raise HTTPException(status_code=502, detail=clean_text(exc))


@app.get("/report-export-link/{code}", dependencies=[Depends(require_api_key)])
def report_export_link(
    code: str,
    request: Request,
    format: str = Query("xlsx", pattern="^(csv|xlsx|pdf)$"),
    mode: str = Query(
        "auto",
        pattern="^(auto|native_export|browser_export|html)$",
        description="Genera el archivo y devuelve un enlace temporal de descarga en JSON.",
    ),
    max_rows: int = Query(500, ge=1, le=50000),
    all_pages: bool = Query(False, description="Si es true, intenta recorrer todas las paginas del informe."),
    max_pages: int = Query(100, ge=1, le=2000, description="Limite de paginas a recorrer cuando all_pages=true."),
    start_date: Optional[str] = Query(None, description="Fecha inicial. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    end_date: Optional[str] = Query(None, description="Fecha final. Acepta YYYY-MM-DD o DD/MM/YYYY."),
    date_from: Optional[str] = Query(None, description="Alias de start_date."),
    date_to: Optional[str] = Query(None, description="Alias de end_date."),
    filters: Optional[str] = Query(None, description="JSON con filtros explicitos por nombre de campo."),
    client: BifarmaClient = Depends(get_client),
) -> JSONResponse:
    try:
        requested_filters = build_requested_filters(request, start_date, end_date, date_from, date_to, filters)
        exported = generate_export_file(
            client=client,
            code=code,
            requested_filters=requested_filters,
            format=format,
            mode=mode,
            max_rows=max_rows,
            all_pages=all_pages,
            max_pages=max_pages,
        )
        stored = store_download_file(
            content=exported["content"],
            filename=exported["filename"],
            media_type=exported["media_type"],
            export_mode=exported["export_mode"],
        )
        base_url = str(request.base_url).rstrip("/")
        download_url = f"{base_url}/download/{stored['file_id']}?token={stored['token']}"
        return JSONResponse(
            {
                "ok": True,
                "version": APP_VERSION,
                "report": exported.get("report"),
                "filename": stored["filename"],
                "format": format,
                "mode": mode,
                "export_mode": stored["export_mode"],
                "size_bytes": stored["size_bytes"],
                "expires_at": stored["expires_at"],
                "download_url": download_url,
                "message": "Archivo generado. Abre download_url en el navegador para descargarlo.",
            }
        )
    except Exception as exc:
        return handle_bifarma_error(exc)


@app.get("/download/{file_id}")
def download_file(file_id: str, token: str = Query(...)) -> StreamingResponse:
    record = get_download_file(file_id, token)
    return StreamingResponse(
        io.BytesIO(record["content"]),
        media_type=record["media_type"],
        headers={
            "Content-Disposition": f'attachment; filename="{record["filename"]}"',
            "X-Bifarma-Export-Mode": record.get("export_mode", ""),
        },
    )
